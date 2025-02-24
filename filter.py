


#this filters out only BOA

# import os

# # Controls
# START_ROW = 2  # Row to start processing (1-based index)
# COLUMN_INDEX = 8  # Column to filter by (1-based index)
# FILTER_CRITERIA = "BANK OF AMERICA, N.A."  # Data to filter for

# # File paths
# DATA_DIR = os.path.join(os.getcwd(), 'data')
# NOT_FINISHED_DIR = os.path.join(DATA_DIR, 'not_finished')
# FINISHED_DIR = os.path.join(DATA_DIR, 'finished')

# INPUT_FILE = os.path.join(NOT_FINISHED_DIR, 'example.xlsx')
# OUTPUT_FILE = os.path.join(FINISHED_DIR, 'example_out.xlsx')

# def process_excel(input_file, output_file, start_row, column_index, filter_criteria):
#     """
#     Filters rows in an XLSX file based on a specific column value.
#     Args:
#         input_file (str): Path to the input XLSX file.
#         output_file (str): Path to save the filtered XLSX file.
#         start_row (int): Row number to start processing from.
#         column_index (int): Column index to filter by (1-based).
#         filter_criteria (str): The value to filter rows.
#     """
#     from openpyxl import load_workbook, Workbook

#     print(f"Loading workbook from {input_file}...")
#     workbook = load_workbook(input_file)
#     sheet = workbook.active

#     # Create a new workbook for filtered data
#     filtered_workbook = Workbook()
#     filtered_sheet = filtered_workbook.active

#     print(f"Filtering rows starting from row {start_row} in column {column_index}...")
#     for row in sheet.iter_rows(min_row=start_row, values_only=True):
#         if row[column_index - 1] != filter_criteria:
#             filtered_sheet.append(row)

#     print(f"Saving filtered data to {output_file}...")
#     os.makedirs(os.path.dirname(output_file), exist_ok=True)
#     filtered_workbook.save(output_file)
#     print("Filtering complete.")

# if __name__ == "__main__":
#     print("Starting Excel filtering process...")
#     process_excel(INPUT_FILE, OUTPUT_FILE, START_ROW, COLUMN_INDEX, FILTER_CRITERIA)
#     print("Excel filtering process completed.")




#  #this filters everything else out and leaves the BOA
# import os

# # Controls
# START_ROW = 2  # Row to start processing (1-based index)
# COLUMN_INDEX = 8 # Column to filter by (1-based index)
# FILTER_CRITERIA = "BANK OF AMERICA, N.A."  # Data to filter for

# # File paths
# DATA_DIR = os.path.join(os.getcwd(), 'data')
# NOT_FINISHED_DIR = os.path.join(DATA_DIR, 'not_finished')
# FINISHED_DIR = os.path.join(DATA_DIR, 'finished')

# INPUT_FILE = os.path.join(NOT_FINISHED_DIR, 'output_file.xlsx')
# OUTPUT_FILE = os.path.join(FINISHED_DIR, 'allBoa_out.xlsx')

# def process_excel(input_file, output_file, start_row, column_index, filter_criteria):
#     """
#     Filters rows in an XLSX file based on a specific column value.
#     Args:
#         input_file (str): Path to the input XLSX file.
#         output_file (str): Path to save the filtered XLSX file.
#         start_row (int): Row number to start processing from.
#         column_index (int): Column index to filter by (1-based).
#         filter_criteria (str): The value to filter rows.
#     """
#     from openpyxl import load_workbook, Workbook

#     print(f"Loading workbook from {input_file}...")
#     workbook = load_workbook(input_file)
#     sheet = workbook.active

#     # Create a new workbook for filtered data
#     filtered_workbook = Workbook()
#     filtered_sheet = filtered_workbook.active

#     print(f"Filtering rows starting from row {start_row} in column {column_index}...")
#     for row in sheet.iter_rows(min_row=start_row, values_only=True):
#         if row[column_index - 1] == filter_criteria:
#             filtered_sheet.append(row)

#     print(f"Saving filtered data to {output_file}...")
#     os.makedirs(os.path.dirname(output_file), exist_ok=True)
#     filtered_workbook.save(output_file)
#     print("Filtering complete.")

# if __name__ == "__main__":
#     print("Starting Excel filtering process...")
#     process_excel(INPUT_FILE, OUTPUT_FILE, START_ROW, COLUMN_INDEX, FILTER_CRITERIA)
#     print("Excel filtering process completed.")




# #filters out by 2 critias (state and bank) on to a new file 

# import os
# from openpyxl import load_workbook, Workbook

# # Controls
# START_ROW = 2  # Row to start processing (1-based index)
# COLUMN_INDEX_1 = 6  # First column to filter by (1-based index)
# COLUMN_INDEX_2 = 8  # Second column to filter by (1-based index)
# FILTER_CRITERIA_1 = "CA"  # First filter criteria
# FILTER_CRITERIA_2 = "BANK OF AMERICA, N.A."  # Second filter criteria

# # File paths
# DATA_DIR = os.path.join(os.getcwd(), 'data')
# NOT_FINISHED_DIR = os.path.join(DATA_DIR, 'not_finished')
# FINISHED_DIR = os.path.join(DATA_DIR, 'finished')

# INPUT_FILE = os.path.join(NOT_FINISHED_DIR, 'output_file.xlsx')
# OUTPUT_FILE = os.path.join(FINISHED_DIR, 'caliBOA.xlsx')

# def process_excel(input_file, output_file, start_row, column_index_1, column_index_2, filter_criteria_1, filter_criteria_2):
#     """
#     Filters rows in an XLSX file based on specific column values.
#     Args:
#         input_file (str): Path to the input XLSX file.
#         output_file (str): Path to save the filtered XLSX file.
#         start_row (int): Row number to start processing from.
#         column_index_1 (int): First column index to filter by (1-based).
#         column_index_2 (int): Second column index to filter by (1-based).
#         filter_criteria_1 (str): The value to filter rows in the first column.
#         filter_criteria_2 (str): The value to filter rows in the second column.
#     """
#     print(f"Loading workbook from {input_file}...")
#     workbook = load_workbook(input_file)
#     sheet = workbook.active

#     # Create a new workbook for filtered data
#     filtered_workbook = Workbook()
#     filtered_sheet = filtered_workbook.active

#     print(f"Filtering rows starting from row {start_row} in columns {column_index_1} and {column_index_2}...")
#     for row in sheet.iter_rows(min_row=start_row, values_only=True):
#         if row[column_index_1 - 1] == filter_criteria_1 and row[column_index_2 - 1] == filter_criteria_2:
#             filtered_sheet.append(row)

#     print(f"Saving filtered data to {output_file}...")
#     os.makedirs(os.path.dirname(output_file), exist_ok=True)
#     filtered_workbook.save(output_file)
#     print("Filtering complete.")

# if __name__ == "__main__":
#     print("Starting Excel filtering process...")
#     process_excel(
#         INPUT_FILE, 
#         OUTPUT_FILE, 
#         START_ROW, 
#         COLUMN_INDEX_1, 
#         COLUMN_INDEX_2, 
#         FILTER_CRITERIA_1, 
#         FILTER_CRITERIA_2
#     )
#     print("Excel filtering process completed.")














import os

# Controls
START_ROW = 2  # Row to start processing (1-based index)
COLUMN_INDEX = 8  # Column to filter by (1-based index)
FILTER_CRITERIA = "PNC BANK"  # Base string to search for, without state

# File paths
DATA_DIR = os.path.join(os.getcwd(), 'data')
NOT_FINISHED_DIR = os.path.join(DATA_DIR, 'not_finished')
FINISHED_DIR = os.path.join(DATA_DIR, 'finished')
INPUT_FILE = os.path.join(NOT_FINISHED_DIR, 'output_file.xlsx')
OUTPUT_FILE = os.path.join(FINISHED_DIR, 'allpnc_out.xlsx')

def process_excel(input_file, output_file, start_row, column_index, filter_criteria):
    """
    Filters rows in an XLSX file based on a partial string match in a specific column.
    
    Args:
        input_file (str): Path to the input XLSX file.
        output_file (str): Path to save the filtered XLSX file.
        start_row (int): Row number to start processing from.
        column_index (int): Column index to filter by (1-based).
        filter_criteria (str): The base string to search for (without state).
    """
    from openpyxl import load_workbook, Workbook
    
    print(f"Loading workbook from {input_file}...")
    workbook = load_workbook(input_file)
    sheet = workbook.active
    
    # Create a new workbook for filtered data
    filtered_workbook = Workbook()
    filtered_sheet = filtered_workbook.active
    
    # Copy header row if starting from row 2 or later
    if start_row > 1:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        filtered_sheet.append(header_row)
    
    print(f"Filtering rows starting from row {start_row} in column {column_index}...")
    matched_count = 0
    
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        cell_value = str(row[column_index - 1] or '')  # Convert to string and handle None
        
        # Check if the filter criteria appears at the start of the cell value
        if cell_value.strip().startswith(filter_criteria):
            filtered_sheet.append(row)
            matched_count += 1
    
    print(f"Found {matched_count} matching rows")
    print(f"Saving filtered data to {output_file}...")
    
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    filtered_workbook.save(output_file)
    print("Filtering complete.")

if __name__ == "__main__":
    print("Starting Excel filtering process...")
    process_excel(INPUT_FILE, OUTPUT_FILE, START_ROW, COLUMN_INDEX, FILTER_CRITERIA)
    print("Excel filtering process completed.")

#     Key changes I made to improve the code:

# Modified the FILTER_CRITERIA to just "PNC CAKE" without the state
# Changed the matching logic to use startswith() instead of exact matching
# Added handling for None values to prevent errors
# Added a counter to show how many matches were found
# Added header row copying to preserve column headers
# Improved string handling with strip() to handle potential whitespace